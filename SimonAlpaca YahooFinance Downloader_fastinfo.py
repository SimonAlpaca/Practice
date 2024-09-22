# -*- coding: utf-8 -*-

import yfinance as yf
import time
from datetime import datetime
import threading
import os
import sys
import tkinter as tk
from ttkbootstrap import Style
from tkinter import ttk
from ctypes import windll
import configparser
import traceback
import openpyxl
import win32com.client

# In[GUI]
class WindowGUI(tk.Frame):
    
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.stock_info = {}
        self.stock_list = []
        self.to_dir = ""
        self.auto = False
        self.stop = False
        self.time_spent = 3
        self.select_range = None
        self.select_col = None
        self.config_quote = None
        self.config_dir = None
        style.configure("primary.TButton", font=("Helvetica", 11,"bold")) # letter button
        
        self.parent.overrideredirect(True)                                  
        self.parent.geometry('540x420+300+150')                                 # window size  
        self.parent.resizable(width=False,height=False)                   # disallow window resize
        self.parent.title("SimonAlpaca Yahoo Finance Downloader")                  # title
        self.parent.withdraw()
        
        self.frame_top1 = ttk.Frame(self.parent,style='Warning.TFrame')             # add frame first
        self.frame_top1.pack(pady=1, side = tk.TOP, fill = "both")
        self.quit_button = ttk.Button(self.frame_top1, text= "\u03A7", width=3, style='primary.Outline.TButton', command = self.quit)
        self.quit_button.pack(side =tk.RIGHT, padx = 10)
        
        self.frame_mid1 = ttk.Frame(self.parent,style='Warning.TFrame')             # add frame first
        self.frame_mid1.pack(pady=1, side = tk.TOP, fill = "both")
        self.frame_mid2 = ttk.Frame(self.parent,style='Warning.TFrame')             # add frame first
        self.frame_mid2.pack(pady=1, side = tk.TOP, fill = "both")
        self.frame_mid3 = ttk.Frame(self.parent,style='Warning.TFrame')             # add frame first
        self.frame_mid3.pack(pady=5, side = tk.TOP, fill = "both")
        self.frame_mid4 = ttk.Frame(self.parent,style='Warning.TFrame')             # add frame first
        self.frame_mid4.pack(pady=8, side = tk.TOP, fill = "both")
        
        self.quote_label = tk.Label(self.frame_mid1, text = "    Input Stock Quote :", fg = "Yellow", font = ("Arial", 10))
        self.quote_label.pack(side = tk.LEFT)
        self.quote_text = tk.Text(self.frame_mid2, height=4, width=65, background='gray25')
        self.quote_text.pack(side=tk.TOP)
        self.dir_label = tk.Label(self.frame_mid3, text = "    Output Directory :", fg = "Yellow", font = ("Arial", 10))
        self.dir_label.pack(side = tk.LEFT)
        self.dir_text = tk.Text(self.frame_mid4, height=1, width=65, background='gray25')
        self.dir_text.pack(side=tk.TOP)
        
        self.frame_bot1 = ttk.Frame(self.parent,style='Warning.TFrame')             # add frame first
        self.frame_bot1.pack(pady=5, side = tk.TOP, fill = "both")
        self.frame_bot2 = ttk.Frame(self.parent,style='Warning.TFrame')             # add frame first
        self.frame_bot2.pack(pady=1, padx=35, side = tk.LEFT)
        self.frame_bot3 = ttk.Frame(self.parent,style='Warning.TFrame')             # add frame first
        self.frame_bot3.pack(pady=5, padx=35, side = tk.RIGHT, fill = "both")
        
        self.progress_label = tk.Label(self.frame_bot1, text = "    Progress :", fg = "Yellow", font = ("Arial", 10))
        self.progress_label.pack(side = tk.LEFT)
        self.progress_text = tk.Text(self.frame_bot2, height=9, width=35, background='gray25')
        self.progress_text.pack(side=tk.LEFT)
        self.progress_scroll = ttk.Scrollbar(self.frame_bot2, orient='vertical', command=self.progress_text.yview)
        self.progress_scroll.pack(side=tk.RIGHT, fill='y')
        self.progress_text['yscrollcommand'] = self.progress_scroll.set
        self.refresh_button = ttk.Button(self.frame_bot3, text= "Refresh", width=30, style='primary.TButton', command=self.refresh)
        self.refresh_button.pack(side=tk.BOTTOM, pady = 10, fill = "both")              # padx means gap of x-axis
        self.auto_button = ttk.Button(self.frame_bot3, text= "Auto", width=30, style='primary.TButton', command= self.auto_refresh)
        self.auto_button.pack(side=tk.BOTTOM, pady = 10, fill = "both")              # padx means gap of x-axis
        self.column_button = ttk.Button(self.frame_bot3, text= "Columns", width=30, style='primary.TButton', command= self.col_con_show)
        self.column_button.pack(side=tk.BOTTOM, pady = 10, fill = "both")              # padx means gap of x-axis
        
        # Window Drag
        self.frame_top1.bind('<Button-1>', self.window_predrag)
        self.frame_top1.bind('<B1-Motion>', self.window_drag)
                
        # Program icon
        try:                                                   
            exe_dir = os.path.split(sys.argv[0])[0]            # sys.argv[0] is the exe path
            ico_path = os.path.join(exe_dir, "Alpaca_ico.ico")
            self.parent.iconbitmap(ico_path)                   # must be parent otherwise can't show program in taskbar   
        
        except:
            pass
        
        self.parent.protocol("WM_DELETE_WINDOW", self.quit)               # when close in taskbar
        
        self.parent.update()
        self.set_appwindow()
        
        self.import_config()
        
        self.col_con = tk.Toplevel(self.parent)
        self.column_controller()
        
    def set_appwindow(self):
        
        GWL_EXSTYLE=-20                                       # for showing program in taskbar
        WS_EX_APPWINDOW=0x00040000
        WS_EX_TOOLWINDOW=0x00000080
        
        hwnd = windll.user32.GetParent(self.parent.winfo_id())
        style = windll.user32.GetWindowLongW(hwnd, GWL_EXSTYLE)
        style = style & ~WS_EX_TOOLWINDOW
        style = style | WS_EX_APPWINDOW
        windll.user32.SetWindowLongW(hwnd, GWL_EXSTYLE, style)
        
        # re-assert the new window style
        self.parent.wm_withdraw()
        self.after(100, self.show_appwindow)
    
    def show_appwindow(self):
        
        self.parent.state("zoomed")                    # fix bug of "can't bring to back before dragging the window"
        self.parent.state("normal")
        
        self.parent.wm_deiconify()
        self.parent.attributes('-topmost', True)
        
        self.parent.update()
        
        time.sleep(0.05)                          # perform better after sleep
        self.parent.attributes('-topmost', False)      # force window to the front but not always

    def window_predrag(self, event):
        
        self.offsetx = event.x
        self.offsety = event.y
        
    def window_drag(self, event):
        
        x = self.parent.winfo_x() + event.x - self.offsetx
        y = self.parent.winfo_y() + event.y - self.offsety
    
        self.parent.geometry('+{x}+{y}'.format(x=x,y=y))
            
    def import_config(self):
        # Import config
        self.config = configparser.ConfigParser()
        
        try:
            exe_dir = os.path.split(sys.argv[0])[0]        # sys.argv[0] is the exe path
            self.config_path = os.path.join(exe_dir, "config_yf_fi.ini")
            self.config.read(self.config_path)
            self.config_dir = self.config.get("YF_Download", "Dir")
            self.config_quote = self.config.get("YF_Download", "Quote")
            self.select_range = self.config.get("YF_Download", "Select_Range").split()
            self.select_col = self.config.get("YF_Download", "Select_Col")

            self.dir_text.insert(tk.INSERT, self.config_dir)
            self.quote_text.insert(tk.INSERT, self.config_quote)
        
        except Exception as e:
            print(e)
            
    def quit(self):
        
        try:
            self.config_file = open(self.config_path, "w")
        
        except PermissionError:
            raise PermissionError("No write permission of the directory") 

        try:
            self.config.add_section("YF_Download")
            
        except configparser.DuplicateSectionError:
            pass                                   # exception if already have config and section
        
        finally:
            quote_get = str(self.quote_text.get(1.0, "end-1c"))
            dir_get = str(self.dir_text.get(1.0, "end-1c"))
            select_get = " ".join(self.select_range)
            self.config.set("YF_Download", "Quote", str(quote_get))
            self.config.set("YF_Download", "Dir", str(dir_get))
            self.config.set("YF_Download", "Select_Range", str(select_get))
            self.config.set("YF_Download", "Select_Col", str(self.select_col))
            self.config.write(self.config_file)
            self.config_file.close()
        
        window.parent.destroy()
        
# In[Download]
    def auto_refresh(self):
        print("auto_refresh")
        
        self.auto = True
        self.auto_button.config(text = "Pause", command=self.pause_refresh)
        self.progress_text.insert("1.0", "  *** Auto Refresh ***\n")
        self.refresh_button.config(state=tk.DISABLED)
        self.refresh()

        while True:
            if self.auto and not self.stop:
                self.parent.update()
                
                sleep_time = max(0, 3 - self.time_spent)
                
                while sleep_time > 0 and self.auto and not self.stop:
                    sleep_time = sleep_time - 0.1
                    time.sleep(0.1)
                    self.parent.update()
                    
                self.refresh()
                
            else:
                break
            
    def pause_refresh(self):
        print("pause_refresh")
        
        self.auto = False
        self.auto_button.config(text = "Auto", command=self.auto_refresh)
        self.progress_text.insert("1.0", "  *** Auto Refresh Stopped ***\n")
        self.refresh_button.config(state=tk.NORMAL)
        self.parent.update()
    
    def refresh(self):
        
        timer = time.perf_counter()
        self.stock_info = {}
        self.stock_list = []
        self.err_dict = {}
        self.is_err = False
        self.progress = ""
        quote_get = str(self.quote_text.get(1.0, "end-1c"))
        self.stock_list = quote_get.split()
        #print(self.stock_list)
        
        self.progress_text.insert("1.0", "  *** Loading ***\n")
        
        self.stop = False
        self.refresh_button.config(text = "Stop Refresh", command=self.stop_refresh)
        self.parent.update()
        
        for stock in self.stock_list:
            globals()['thread_ticker_%s' % stock] = threading.Thread(target = self.get_info, args=(stock,))
            globals()['thread_ticker_%s' % stock].start()
        
        self.parent.update()
        if self.stop:
            self.stop_finish()
            
        for stock in self.stock_list:
            self.parent.update()
            globals()['thread_ticker_%s' % stock].join(timeout = 5)
            
        self.parent.update()
        if self.stop:
            self.stop_finish()
        
        for stock in self.stock_list:
            self.parent.update()
            if self.stop:
                self.stop_finish()
            self.get_download(stock)     # yf.download method does not support multi-thread   
            
        self.write_print()
        
        self.create_excel()                          
        self.direct_write_excel()                    # refresh opened excel directly, but can't work if working on the excel
        
        if self.is_err:
            self.progress_text.delete('1.0', '2.0')
            err_cantwrite()
        
        else:
            self.progress_text.delete('1.0', '2.0')
            self.time_spent = round(time.perf_counter() - timer,3)
            print("Time spent : %f\n" %self.time_spent)
            
            now = datetime.now()
            current_time = now.strftime("%H:%M:%S")
            
            self.progress_text.insert("1.0", "%s Success; Time Spent : %.3fs\n" %(current_time, self.time_spent))
            
        self.refresh_button.config(text = "Refresh", command=self.refresh)
        self.parent.update()
        
    def stop_refresh(self):
        print("stop_refresh")
        
        self.stop = True
        self.parent.update()
        
        self.stop_finish()
        
    def stop_finish(self):
        print("stop_finish")
        
        self.refresh_button.config(text = "Refresh", command=self.refresh)
        self.parent.update()
        self.progress_text.insert("1.0", "  *** Refresh Stopped ***\n")
        
        raise Exception("Refresh Stopped")

    def get_info(self, stock):
        
        if stock != "***":
            if stock.isnumeric():
                stock_code = stock.rjust(4,"0") + ".hk"
            
            else:
                stock_code = stock
            
            ticker = yf.Ticker(stock_code)
            self.stock_info[stock] = ticker.fast_info
                
            # print(ticker.fast_info)

    def get_download(self, stock):
        
        if stock != "***":
            if "lastPrice" not in self.stock_info[stock]:
                download_txt = yf.download(stock, period="1D") # use another method if there is no currentPrice in ticket.fast_info
                close_price = download_txt["Close"].tolist()[0]
                self.stock_info[stock]["lastPrice"] = close_price   # insert close price
            
    def write_print(self):
        
        self.to_dir = str(self.dir_text.get(1.0, "end-1c"))
        
        output_path = os.path.join(self.to_dir,"yahoofin_data.txt")
        
        try:
            output_file = open(output_path, "w", encoding ="utf-8")   # create txt
        
        except FileNotFoundError:
            err_filenotfound()
            
        output_file.flush()
        
        # print("Stock\t\t\t Quote\t\t Price\t Change\t %\t\t Vol")
        # output_file.write("Stock\tQuote\tPrice\tChange\t%\tVolume\n")
        
        column_list = self.select_col.split()

        for i in range(len(column_list)):
            if i == len(column_list) - 1:
                print(column_list[i])
                output_file.write("%s" %column_list[i])
                
            else:
                print(column_list[i], end="\t")
                output_file.write("%s\t" %column_list[i])
        
        print()
        output_file.write("\n")
        
        for stock in self.stock_list:                         # write txt
            if stock != "***":
                #print(self.stock_info[stock])
                try:
                    self.progress = stock
                    for i in range(len(column_list)):
                        try:
                            if column_list[i] == "quote":
                                col_data = self.quote_col(stock)
                            
                            elif column_list[i] == "stock_price_change":
                                col_data = self.stock_price_change(stock)
                            
                            elif column_list[i] == "stock_price_change_pct":
                                col_data = self.stock_price_change_pct(stock)
                                    
                            else:
                                col_data = self.stock_info[stock][column_list[i]]
                        
                        except KeyError:
                            if i == len(column_list) - 1:
                                print("")
                                output_file.write("")
                                
                            else:
                                print("\t")
                                output_file.write("\t")
                            
                        else:
                            if i == len(column_list) - 1:
                                print(col_data)
                                output_file.write("%s" % col_data)
                                
                            else:
                                print(col_data, end="\t")
                                output_file.write("%s\t" % col_data)

                    print()
                    output_file.write("\n")
                    output_file.flush()
                
                except KeyError:
                    self.err_dict[self.progress] = traceback.format_exc()
                    self.is_err = True
                    
            else:
                print("******")
                output_file.write("******\n")
                
                output_file.flush()
    
        output_file.close()
    
    def create_excel(self):
        wb = openpyxl.Workbook()
        
        excel_name = "yahoofin_data.xlsx"
        sheet = wb.create_sheet("YahooFinance", 0)
        
        try:
            wb.save(excel_name)
        
        except PermissionError:
            print("PermissionError")
            wb = openpyxl.load_workbook(excel_name)
            sheet = wb.active
    
    def direct_write_excel(self):
        # ExcelApp = win32com.client.gencache.EnsureDispatch('Excel.Application')
        
        output_path = os.path.join(self.to_dir,"yahoofin_data.xlsx")
        
        ExcelApp = win32com.client.GetActiveObject("Excel.Application")
        
        ExcelApp.Visible = True
        
        # Open the desired workbook
        workbook = ExcelApp.Workbooks.Open(output_path)
        
        worksheet = workbook.Worksheets("YahooFinance")
        
        start_row = 1
        start_col = 1
        
        self.to_dir = str(self.dir_text.get(1.0, "end-1c"))
        
        column_list = self.select_col.split()

        for i in range(len(column_list)):
            worksheet.Cells(start_row,start_col + i).Value = column_list[i]
        
        j = 0
        for stock in self.stock_list:                         # write stock value
            j = j + 1
            if stock != "***":
                #print(self.stock_info[stock])
                try:
                    self.progress = stock
                    for i in range(len(column_list)):
                        try:
                            
                            if column_list[i] == "quote":
                                col_data = self.quote_col(stock)
                                
                            elif column_list[i] == "stock_price_change":
                                col_data = self.stock_price_change(stock)
                            
                            elif column_list[i] == "stock_price_change_pct":
                                col_data = self.stock_price_change_pct(stock)
                                
                            else:
                                col_data = self.stock_info[stock][column_list[i]]
                        
                        except KeyError:
                            try:
                                worksheet.Cells(start_row + j, start_col + i).Value = ""
                            
                            except:
                                print("can't write")
        
                            
                        else:
                            try:
                                worksheet.Cells(start_row + j, start_col + i).Value = col_data
                            
                            except:
                                print("can't write")
                
                except KeyError:
                    self.err_dict[self.progress] = traceback.format_exc()
                    self.is_err = True
                    
            else:
                try:
                    worksheet.Cells(start_row + j, start_col).Value = "******"
                    
                except:
                    print("can't write")
             
# In[Column Manager]
    def column_controller(self):
        
        s = ttk.Style()
        s.configure("primary.TButton", font=("Helvetica", 11,"bold")) # letter button
        
        self.col_con.resizable(False,False)
        self.col_con.overrideredirect(1)
        self.col_con.geometry('600x260+830+200')                                 # column controller size 
        
        self.frame_colcon0 = ttk.Frame(self.col_con,style='Warning.TFrame', height=20)             # add frame first
        self.frame_colcon0.pack(pady=0, side = tk.TOP, fill = "both")
        self.frame_colcon1 = ttk.Frame(self.col_con,style='Warning.TFrame')             # add frame first
        self.frame_colcon1.pack(pady=5, side = tk.TOP, fill = "both")
        self.frame_colcon2 = ttk.Frame(self.col_con,style='Warning.TFrame')             # add frame first
        self.frame_colcon2.pack(pady=1, padx=25, side = tk.TOP)
        self.frame_colcon2_1 = ttk.Frame(self.frame_colcon2,style='Warning.TFrame')             # add frame first
        self.frame_colcon2_1.pack(pady=10, padx=10, side = tk.LEFT, fill = "both")
        self.frame_colcon2_2 = ttk.Frame(self.frame_colcon2,style='Warning.TFrame')             # add frame first
        self.frame_colcon2_2.pack(pady=10, padx=150, side = tk.RIGHT, fill = "both")
        self.frame_colcon4 = ttk.Frame(self.col_con,style='Warning.TFrame')             # add frame first
        self.frame_colcon4.pack(pady=10, padx=10, side = tk.TOP, fill = "both")
        
        self.col_text = tk.Text(self.frame_colcon1, height=5, width=80, background='gray25')
        self.col_text.pack(side=tk.TOP)
        
        if self.select_col is None:
            self.select_col = "quote lastPrice stock_price_change stock_price_change_pct open dayHigh dayLow "
        
        if self.config_quote is None:
            self.quote_text.insert(tk.INSERT, "QQQ TQQQ TSLA ^VIX BTC-USD ETH-USD GC=F BZ=F CL=F GBPHKD=X USDHKD=X")

        if self.config_dir is None:
            exe_dir = os.path.split(sys.argv[0])[0]        # sys.argv[0] is the exe path
            self.dir_text.insert(tk.INSERT, exe_dir)
        
        self.col_text.insert("1.0", "%s " %self.select_col)
            
        self.cancel_button = ttk.Button(self.frame_colcon4, text= "Cancel", width=20, style='primary.TButton', command = self.col_con_cancel)
        self.cancel_button.pack(side =tk.RIGHT, padx = 10)
        
        self.ok_button = ttk.Button(self.frame_colcon4, text= "Save and Close", width=30, style='primary.TButton', command = self.col_con_save)
        self.ok_button.pack(side =tk.LEFT, padx = 10)
        
        if self.select_range is None:
            print("default select_range")
            ticker = yf.Ticker("0001.hk")
            custom_col = ["quote", "stock_price_change", "stock_price_change_pct"]
            self.select_range = list(ticker.fast_info.keys()) + custom_col
            self.select_range.remove("marketCap")
            self.select_range.sort()
            
        select_var = tk.StringVar(self)
        self.select_scroll = ttk.OptionMenu(self.frame_colcon2_1, select_var, self.select_range[0], *self.select_range, style='primary.TButton',command = lambda e: self.col_text.insert(tk.END, "%s " %select_var.get()))
        #self.select_scroll["highlightthickness"] = 0                  # remove the option boundary
        self.select_scroll.pack(side =tk.RIGHT, pady = 10)
        
        self.commonlabel = ttk.Label(self.frame_colcon2_1, text="Add Columns :         ", font = ("Arial", 11, "bold"))
        self.commonlabel.pack(side = tk.LEFT, pady = 10)
        
        self.col_con.attributes('-topmost', False)
        self.col_con.withdraw()

        # Gif Controller Drag
        self.col_con_offsetx = 0
        self.col_con_offsety = 0
        self.frame_colcon0.bind('<Button-1>', self.col_con_predrag)
        self.frame_colcon0.bind('<B1-Motion>', self.col_con_drag)
    
    def col_con_predrag(self, event):
        
        self.col_con_offsetx = event.x
        self.col_con_offsety = event.y
        
    def col_con_drag(self, event):
        
        x = self.col_con.winfo_x() + event.x - self.col_con_offsetx
        y = self.col_con.winfo_y() + event.y - self.col_con_offsety
    
        self.col_con.geometry('+{x}+{y}'.format(x=x,y=y))
    
    def col_con_show(self):
        
        self.col_con.deiconify()
        self.col_con.attributes('-topmost', True)
    
    def col_con_cancel(self):
        
        self.col_text.delete("1.0", tk.END)
        self.col_text.insert("1.0", "%s " %self.select_col)
        
        self.col_con.attributes('-topmost', False)
        self.col_con.withdraw()
    
    def col_con_save(self):
        
        self.select_col = str(self.col_text.get("1.0", tk.END))
        
        self.col_con.attributes('-topmost', False)
        self.col_con.withdraw()
    
# In[Custom Column]

    def quote_col(self,stock):
        
        return stock

    def stock_price_change(self, stock):
        
        stock_price = self.stock_info[stock]["lastPrice"]
        stock_preprice = self.stock_info[stock]["previousClose"]
        
        return round(stock_price - stock_preprice, 4)
        
    def stock_price_change_pct(self,stock):
        
        stock_price = self.stock_info[stock]["lastPrice"]
        stock_preprice = self.stock_info[stock]["previousClose"]
        stock_price_change = stock_price - stock_preprice
        
        return round(stock_price_change / stock_preprice, 4)
    
# In[Exception]

def err_filenotfound():
    
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    window.progress_text.insert("1.0", "%s Fail;No such file or directory" %current_time)
    raise FileNotFoundError("No such file or directory")

def err_cantwrite():
    
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    
    list_keys = list(window.err_dict.keys())
    list_values = list(window.err_dict.values())
    
    for i in range(len(list_keys)):
        print("Cant write : %s" % list_keys[i])
        print(list_values[i])
        window.progress_text.insert("1.0", "%s" % list_values[i])
        window.progress_text.insert("1.0", "%s\n" % list_keys[i])
    
    window.progress_text.insert("1.0", "%s Fail;\n" %(current_time))
    
# In[Initial]

if __name__ == "__main__":

    style = Style(theme='darkly')                             # need manual modification of theme file
    style_master = style.master                                         # create window by ttk

    window = WindowGUI(style_master)
    
    window.mainloop()        # must add at the end to make it run
    
    os._exit(1)

# In[things to do]

'''
Functionality:

Optimization:
    
Bug:
    
'''

# In[convert to exe]

'''
https://ithelp.ithome.com.tw/articles/10231524

1) Open Windows PowerShell
2) (One-time) Install Python (go to python website to install)
3) (One-time) Install module including yfinance, datetime, openpyxl
4) open auto-py-to-exe
5) 
-- onedir
-- noconfirm
-- icon
-- collect all: ttkbootstrap

pyinstaller --noconfirm --onedir --windowed --icon "C:/Users/simon/Practice/alpaca_ico.ico" --no-embed-manifest --collect-all "ttkbootstrap"  "C:/Users/simon/Practice/SimonAlpaca YahooFinance Downloader.py"

6) replace ttkbootstrap folder
'''

